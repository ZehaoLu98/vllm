import re, os, glob
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from util import parse_human_readable_int

# --- Configuration ---
LOGS_DIR = "/home/ubuntu/Inference/vllm/vllm_profile/experiment_results/experiment_20260217_210309"
LOG_PATTERN = "batch_*_tokens_*_output.log"
OUTPUT_FILE = "/home/ubuntu/Inference/vllm/vllm_profile/experiment_results/experiment_20260217_210309/performance_metrics.xlsx"

# --- Discover and parse log files ---
log_files = sorted(glob.glob(os.path.join(LOGS_DIR, LOG_PATTERN)),
                   key=lambda f: int(re.search(r'batch_(\d+)_tokens', f).group(1)))

data = []
failed = []

for path in log_files:
    # Extract batch size and token budget from filename
    filename_match = re.search(r'batch_(\d+)_tokens_([^_]+)_output', path)
    bs = int(filename_match.group(1))
    token_budget_str = filename_match.group(2)
    token_budget = parse_human_readable_int(token_budget_str)
    with open(path, 'r') as f:
        content = f.read()

    m_ttft = re.search(
        r'Time To First Token.*?Average:\s+([\d.]+).*?Min:\s+([\d.]+).*?Max:\s+([\d.]+)',
        content, re.DOTALL)
    m_tpot = re.search(
        r'Time Per Output Token.*?Average:\s+([\d.]+).*?Min:\s+([\d.]+).*?Max:\s+([\d.]+)',
        content, re.DOTALL)

    if not m_ttft or not m_tpot:
        failed.append((bs, token_budget))
        continue

    prompt_tp = [float(x) for x in re.findall(r'Avg prompt throughput:\s+([\d.]+)', content)]
    output_tp = [float(x) for x in re.findall(r'Avg generation throughput:\s+([\d.]+)', content)]

    data.append({
        'batch_size': bs,
        'token_budget': token_budget,
        'ttft_avg': float(m_ttft.group(1)), 'ttft_min': float(m_ttft.group(2)), 'ttft_max': float(m_ttft.group(3)),
        'tpot_avg': float(m_tpot.group(1)), 'tpot_min': float(m_tpot.group(2)), 'tpot_max': float(m_tpot.group(3)),
        'prompt_tp': sum(prompt_tp) / len(prompt_tp) if prompt_tp else 0,
        'output_tp': sum(output_tp) / len(output_tp) if output_tp else 0,
    })

if failed:
    print(f"OOM/failed batch sizes (excluded): {failed}")

# --- Create Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Performance Metrics"

headers = ['Batch Size', 'Token Budget', 'TTFT Avg (s)', 'TTFT Min (s)', 'TTFT Max (s)',
           'TPOT Avg (s)', 'TPOT Min (s)', 'TPOT Max (s)',
           'Prompt Throughput (tok/s)', 'Output Throughput (tok/s)']

hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill('solid', fgColor='4472C4')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Arial', size=11)
data_align = Alignment(horizontal='center')
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
alt_fill = PatternFill('solid', fgColor='D9E2F3')

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

for ri, d in enumerate(data, 2):
    vals = [d['batch_size'], d['token_budget'], d['ttft_avg'], d['ttft_min'], d['ttft_max'],
            d['tpot_avg'], d['tpot_min'], d['tpot_max'], d['prompt_tp'], d['output_tp']]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=ri, column=col, value=round(v, 4) if isinstance(v, float) else v)
        c.font, c.alignment, c.border = data_font, data_align, border
        if col <= 2:   c.number_format = '#,##0'
        elif col <= 8: c.number_format = '0.0000'
        else:          c.number_format = '#,##0.0'
        if ri % 2 == 0:
            c.fill = alt_fill

# Mark OOM rows
for i, (bs, token_budget) in enumerate(failed):
    ri = len(data) + 2 + i
    c = ws.cell(row=ri, column=1, value=bs)
    c.font, c.alignment, c.border = data_font, data_align, border
    c.number_format = '#,##0'
    c2 = ws.cell(row=ri, column=2, value=token_budget)
    c2.font, c2.alignment, c2.border = data_font, data_align, border
    c2.number_format = '#,##0'
    oom = ws.cell(row=ri, column=3, value='OOM - CUDA out of memory')
    oom.font = Font(name='Arial', size=11, color='FF0000', italic=True)
    oom.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=ri, start_column=3, end_row=ri, end_column=10)

col_widths = [12, 14, 14, 14, 14, 14, 14, 14, 24, 24]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 35
ws.auto_filter.ref = f"A1:J{len(data) + 1}"
ws.freeze_panes = 'A2'

wb.save(OUTPUT_FILE)
print(f"Saved: {OUTPUT_FILE} ({len(data)} rows + {len(failed)} OOM)")
for d in data:
    print(f"  Batch {d['batch_size']:>5} TokenBudget {d['token_budget']:>6}: TTFT={d['ttft_avg']:.2f}/{d['ttft_min']:.2f}/{d['ttft_max']:.2f}"
          f"  TPOT={d['tpot_avg']:.4f}/{d['tpot_min']:.4f}/{d['tpot_max']:.4f}"
          f"  Prompt={d['prompt_tp']:.1f}  Output={d['output_tp']:.1f}")